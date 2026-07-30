#!/usr/bin/env python3
"""Normalize the Norwegian Red List 2021 Excel workbook.

The workbook is downloaded manually from
https://lister.artsdatabanken.no/rodlisteforarter/2021/ (source: Artsdatabanken).
This script reads the workbook in bounded streaming mode, resolves columns by
normalized header name (not fixed column position), validates categories and
areas against the known 2021 taxonomy, and emits a deterministic
``assessments.jsonl`` plus a ``report.json`` describing the input's provenance
and the diagnostics found.

Identity rule — the normalizer does NOT allocate or resolve
``sporely_taxon_id``. Every assessed row carries the workbook's own
``Vitenskapelig navn id`` under the ``artsnavnebase_scientific_name_id``
namespace; the shared compiler resolves that identifier against existing
NorTaxa ``scientificNameID`` bindings.

Output determinism — two runs against identical workbook bytes produce
byte-identical ``assessments.jsonl`` and ``report.json``. No timestamps are
written into the report; provenance is bound by SHA-256.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import sys
from pathlib import Path
from typing import Iterator

# Delayed import so `--help` works without openpyxl installed.


SOURCE_SYSTEM = "artsdatabanken_redlist"
SOURCE_RELEASE = "2021"
SOURCE_URL = "https://lister.artsdatabanken.no/rodlisteforarter/2021/"
ASSESSED_NAME_SOURCE = "artsdatabanken"
ASSESSED_NAME_NAMESPACE = "artsnavnebase_scientific_name_id"
ASSESSMENTS_SHEET = "Vurderinger"

# The 2021 Norwegian Red List uses these nine assessment categories.
# See https://artsdatabanken.no/rodlisteforarter2021/hjelp/kategoriene.
ALLOWED_CATEGORIES = frozenset({"RE", "CR", "EN", "VU", "NT", "DD", "LC", "NA", "NE"})

# The workbook assesses two areas separately. They are never merged.
ALLOWED_AREAS = frozenset({"Norge", "Svalbard"})

# Norwegian taxonomic level → canonical rank whitelist. Anything outside
# this whitelist is stored as NULL rather than guessed.
RANK_MAP = {
    "art": "species",
    "underart": "subspecies",
    "varietet": "variety",
    "form": "form",
    "slekt": "genus",
    "aggregat": "aggregate",
}

# Header → normalized-field mapping. Keys are matched case-insensitively after
# whitespace collapse. The workbook may add columns in later releases; only
# the required set must be present.
REQUIRED_HEADERS = {
    "id for vurderingen": "assessment_id",
    "vurderingsområde": "assessment_area",
    "vitenskapelig navn id": "assessed_name_id",
    "vitenskapelig navn": "scientific_name_snapshot",
    "kategori 2021": "category_raw",
}
OPTIONAL_HEADERS = {
    "autor": "authorship_snapshot",
    "populærnavn": "common_name_snapshot",
    "ekspertkomité": "expert_group",
    "artsgruppe": "species_group",
    "taksonomisk nivå": "taxonomic_level_raw",
    "kriterier 2021": "criteria",
    "begrunnelse nedgradering av kategori": "downgrade_rationale",
    "årstall for siste revisjon": "assessment_year_raw",
}


class RedlistNormalizeError(Exception):
    """Raised when the workbook cannot be normalized deterministically."""


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def _resolve_headers(header_row: tuple) -> dict[str, int]:
    """Build a {field_name: column_index} map, failing on required-column
    misses and on ambiguous duplicates."""
    seen: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        key = _normalize_header(raw)
        if not key:
            continue
        for lookup in (REQUIRED_HEADERS, OPTIONAL_HEADERS):
            if key in lookup:
                field = lookup[key]
                if field in seen:
                    raise RedlistNormalizeError(
                        f"duplicate header for field {field!r}: column {seen[field]} "
                        f"and column {idx}"
                    )
                seen[field] = idx
    missing = [h for h in REQUIRED_HEADERS.values() if h not in seen]
    if missing:
        raise RedlistNormalizeError(
            f"missing required workbook columns: {sorted(missing)!r}"
        )
    return seen


def _classify_category(raw: str) -> tuple[str, str, bool]:
    """Return (category_raw, category_code, is_downgraded).

    Downgrade convention in the 2021 Norwegian Red List uses a trailing
    degree sign (``°``) on the category code. That marker is preserved in
    ``category_raw`` and stripped for ``category_code``.
    """
    if raw is None:
        raise RedlistNormalizeError("category value is missing")
    text = str(raw).strip()
    if not text:
        raise RedlistNormalizeError("category value is empty")
    is_downgraded = text.endswith("°")
    code = text[:-1] if is_downgraded else text
    if code not in ALLOWED_CATEGORIES:
        raise RedlistNormalizeError(
            f"unknown category value: {text!r} — allowed: "
            f"{sorted(ALLOWED_CATEGORIES)!r}"
        )
    return text, code, is_downgraded


def _classify_rank(raw) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip().lower()
    if not text:
        return None
    return RANK_MAP.get(text)


def _clean_optional_text(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return text


def _assessment_url(assessment_id: str) -> str:
    return f"https://artsdatabanken.no/lister/rodlisteforarter/2021/{assessment_id}"


def _iter_workbook_rows(
    workbook_path: Path,
) -> Iterator[tuple[dict[str, int], tuple]]:
    """Yield ``(headers, row)`` for every data row of the assessments sheet."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RedlistNormalizeError(
            "openpyxl is required. Install with: pip install openpyxl"
        ) from exc
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if ASSESSMENTS_SHEET not in wb.sheetnames:
            raise RedlistNormalizeError(
                f"workbook is missing expected sheet {ASSESSMENTS_SHEET!r}; "
                f"sheets found: {wb.sheetnames!r}"
            )
        ws = wb[ASSESSMENTS_SHEET]
        iterator = ws.iter_rows(values_only=True)
        try:
            header_row = next(iterator)
        except StopIteration as exc:
            raise RedlistNormalizeError("assessments sheet is empty") from exc
        headers = _resolve_headers(header_row)
        for row in iterator:
            # Skip fully-blank trailing rows that Excel sometimes emits.
            if not any(cell not in (None, "") for cell in row):
                continue
            yield headers, row
    finally:
        wb.close()


def _row_value(row: tuple, headers: dict[str, int], field: str):
    idx = headers.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def normalize(*, input_path: Path, output_dir: Path) -> dict:
    """Normalize the workbook. Returns the report dict; also writes it to
    ``output_dir/report.json`` and the assessments to
    ``output_dir/assessments.jsonl``.
    """
    if not input_path.exists():
        raise RedlistNormalizeError(f"workbook not found: {input_path}")
    if output_dir.exists():
        if any(output_dir.iterdir()):
            raise RedlistNormalizeError(
                f"output directory is not empty: {output_dir}"
            )
    else:
        output_dir.mkdir(parents=True, exist_ok=False)

    workbook_sha = _sha256_file(input_path)

    assessments: list[dict] = []
    diagnostics = {
        "duplicate_assessment_ids": [],
        "duplicate_name_area_keys": [],
        "invalid_rows": [],
        "unknown_areas": [],
    }
    seen_assessment_ids: set[str] = set()
    seen_name_area_keys: set[tuple[str, str, str, str]] = set()
    category_counts: dict[str, int] = {}
    area_counts: dict[str, int] = {}
    row_number = 1  # header is row 1
    resolved_headers: dict[str, int] | None = None

    for headers, row in _iter_workbook_rows(input_path):
        row_number += 1
        resolved_headers = headers
        try:
            assessment_id = _row_value(row, headers, "assessment_id")
            if assessment_id is None or str(assessment_id).strip() == "":
                raise RedlistNormalizeError("assessment_id is blank")
            assessment_id = str(assessment_id).strip()

            area = _row_value(row, headers, "assessment_area")
            if area is None or str(area).strip() == "":
                raise RedlistNormalizeError("assessment_area is blank")
            area = str(area).strip()
            if area not in ALLOWED_AREAS:
                diagnostics["unknown_areas"].append(
                    {"row": row_number, "area": area}
                )
                raise RedlistNormalizeError(
                    f"unknown assessment area: {area!r}"
                )

            assessed_name_id = _row_value(row, headers, "assessed_name_id")
            if assessed_name_id is None or str(assessed_name_id).strip() == "":
                raise RedlistNormalizeError("assessed_name_id is blank")
            assessed_name_id = str(assessed_name_id).strip()

            scientific_name = _row_value(row, headers, "scientific_name_snapshot")
            if scientific_name is None or str(scientific_name).strip() == "":
                raise RedlistNormalizeError("scientific_name_snapshot is blank")
            scientific_name = str(scientific_name).strip()

            raw_category = _row_value(row, headers, "category_raw")
            category_raw, category_code, downgraded = _classify_category(raw_category)

            if assessment_id in seen_assessment_ids:
                diagnostics["duplicate_assessment_ids"].append(
                    {"row": row_number, "assessment_id": assessment_id}
                )
                raise RedlistNormalizeError(
                    f"duplicate assessment_id: {assessment_id}"
                )
            seen_assessment_ids.add(assessment_id)

            name_area_key = (
                ASSESSED_NAME_SOURCE,
                ASSESSED_NAME_NAMESPACE,
                assessed_name_id,
                area,
            )
            if name_area_key in seen_name_area_keys:
                diagnostics["duplicate_name_area_keys"].append(
                    {"row": row_number, "key": list(name_area_key)}
                )
                raise RedlistNormalizeError(
                    f"duplicate (name-id, area) key: {name_area_key!r}"
                )
            seen_name_area_keys.add(name_area_key)

            rank = _classify_rank(_row_value(row, headers, "taxonomic_level_raw"))

            record = {
                "source_system": SOURCE_SYSTEM,
                "source_release": SOURCE_RELEASE,
                "assessment_id": assessment_id,
                "assessment_area": area,
                "assessed_name_source": ASSESSED_NAME_SOURCE,
                "assessed_name_namespace": ASSESSED_NAME_NAMESPACE,
                "assessed_name_id": assessed_name_id,
                "scientific_name_snapshot": scientific_name,
                "authorship_snapshot": _clean_optional_text(
                    _row_value(row, headers, "authorship_snapshot")
                ),
                "common_name_snapshot": _clean_optional_text(
                    _row_value(row, headers, "common_name_snapshot")
                ),
                "taxon_rank_snapshot": rank,
                "taxonomic_level_raw": _clean_optional_text(
                    _row_value(row, headers, "taxonomic_level_raw")
                ),
                "expert_group": _clean_optional_text(
                    _row_value(row, headers, "expert_group")
                ),
                "species_group": _clean_optional_text(
                    _row_value(row, headers, "species_group")
                ),
                "category_raw": category_raw,
                "category_code": category_code,
                "category_is_downgraded": downgraded,
                "criteria": _clean_optional_text(
                    _row_value(row, headers, "criteria")
                ),
                "downgrade_rationale": _clean_optional_text(
                    _row_value(row, headers, "downgrade_rationale")
                ),
                "assessment_url": _assessment_url(assessment_id),
            }
            assessments.append(record)
            category_counts[category_code] = category_counts.get(category_code, 0) + 1
            area_counts[area] = area_counts.get(area, 0) + 1
        except RedlistNormalizeError as exc:
            diagnostics["invalid_rows"].append(
                {"row": row_number, "error": str(exc)}
            )
            raise

    # Deterministic ordering: by (area, assessment_id numeric-then-lex).
    def sort_key(r: dict) -> tuple:
        try:
            aid_num = (0, int(r["assessment_id"]))
        except ValueError:
            aid_num = (1, r["assessment_id"])
        return (r["assessment_area"], aid_num, r["assessment_id"])

    assessments.sort(key=sort_key)

    assessments_path = output_dir / "assessments.jsonl"
    with assessments_path.open("w", encoding="utf-8") as handle:
        for record in assessments:
            handle.write(json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n")

    header_map = {
        field: (col_idx, _column_letter(col_idx))
        for field, col_idx in sorted((resolved_headers or {}).items())
    }
    report = {
        "source_system": SOURCE_SYSTEM,
        "source_release": SOURCE_RELEASE,
        "source_url": SOURCE_URL,
        "input_filename": input_path.name,
        "input_sha256": workbook_sha,
        "sheet_name": ASSESSMENTS_SHEET,
        "header_map": header_map,
        "row_count": len(assessments),
        "category_counts": dict(sorted(category_counts.items())),
        "area_counts": dict(sorted(area_counts.items())),
        "downgraded_count": sum(
            1 for r in assessments if r["category_is_downgraded"]
        ),
        "assessments_output": {
            "name": assessments_path.name,
            "sha256": _sha256_file(assessments_path),
            "bytes": assessments_path.stat().st_size,
        },
        "allowed_categories": sorted(ALLOWED_CATEGORIES),
        "allowed_areas": sorted(ALLOWED_AREAS),
    }
    report_path = output_dir / "report.json"
    report_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return report


def _column_letter(idx: int) -> str:
    """0-based column index → Excel letters (A, B, ..., AA, AB ...)."""
    n = idx + 1
    out = ""
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(ord("A") + rem) + out
    return out


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True,
                        help="path to the manually downloaded red list workbook")
    parser.add_argument("--output", type=Path, required=True,
                        help="destination directory for normalized outputs "
                             "(must not exist or must be empty)")
    return parser


def main(argv=None) -> int:
    args = build_parser().parse_args(list(argv) if argv is not None else None)
    try:
        report = normalize(input_path=args.input, output_dir=args.output)
    except RedlistNormalizeError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 2
    print(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
